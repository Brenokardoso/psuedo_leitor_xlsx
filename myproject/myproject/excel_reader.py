from openpyxl import load_workbook


# carregando arquivo xlsx
workbook = load_workbook("myproject/chestarc/6.1-a-Indenização de Tansporte Junho de 2023.xlsx")

# seleciona uma planilha
ws = workbook.active

max_row = ws.max_row
start_data = False
end_date = False
check_file = False
line = 0


for row in ws.iter_rows():
    line += 1
    if (row[0] == "(a)"):
        start_data = True
        check_file = True
    elif start_data is True:
        if isinstance(row[0], str):
            if row[0][0:5] == 'Fonte':
                start_data = False
                break
            if all(value is None for value in row ):
                continue
            












# # itera sobre ela
# for rows in ws.iter_rows():
#     for cel in rows:
#         if cel.value is not None:
#             print(f"O valor das celulas = {cel.value}")
#             print("\n")

#         else:
#             continue





